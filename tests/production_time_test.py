
from pathlib import Path

import pytest
import openpyxl as op

from moduler.toolmonitor_data.production_time import production_time


def test_productiontime():

    test_database = Path('./tests/test_data/test_pt_database.xlsx')
    production_time(test_database, '999999', 3600)
    production_time(test_database, '999998', 3601)
    production_time(test_database, '999997', 3599)

    assert test_database.is_file() is True

    workbook = op.load_workbook(test_database)
    worksheet = workbook['Results']

    data = {worksheet.cell(row=2, column=1).value: worksheet.cell(row=2, column=2).value,
            worksheet.cell(row=3, column=1).value: worksheet.cell(row=3, column=2).value,
            worksheet.cell(row=4, column=1).value: worksheet.cell(row=4, column=2).value}

    assert '999999' in data
    assert '999998' in data
    assert '999997' in data

    assert data['999999'] == '1:0:0'
    assert data['999998'] == '1:0:1'
    assert data['999997'] == '0:59:59'


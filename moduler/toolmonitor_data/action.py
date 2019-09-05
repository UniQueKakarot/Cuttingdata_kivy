from pathlib import Path

import openpyxl as op

#TODO
# Calculate how many piecec I can machine before I have to replace any tool
# How much time does it take to produce this product? "Tidskalkyle"


def check_database(database_path):
    """ Generates the excel database if it dosent exist """
    if not database_path.is_file():
        workbook = op.Workbook()

        worksheet1 = workbook.active
        worksheet1.title = "Tidskalkyle"

        workbook.create_sheet("Unused Tools")

        workbook.save(database_path)


def tools_unused(raw_database_path, database_path):

    # path to the databse
    check_database(database_path)

    # load the raw data database
    raw_database = op.load_workbook(raw_database_path)
    if 'Old Data' in raw_database.sheetnames:
        raw_worksheet1 = raw_database['Old Data']
        raw_worksheet2 = raw_database['New Data']

        # load the results database
        database = op.load_workbook(database_path)
        result = database['Unused Tools']

        # loop through the raw data database and compare new and old tool time data to establish
        # what tools was used and not
        row = 2
        unused_tools = []
        while raw_worksheet1.cell(row=row, column=5).value is not None:

            if raw_worksheet1.cell(row=row, column=5).value == raw_worksheet2.cell(row=row, column=5).value:
                unused_tools.append(raw_worksheet1.cell(row=row, column=2).value)

            row += 1

        # loop through the result database to extract any "unused" tools already registrated
        recorded_unused_tools = []
        row = 1
        while result.cell(row=row, column=1).value is not None:
            recorded_unused_tools.append(result.cell(row=row, column=1).value)
            row += 1

        # check if anything is written in the result database
        # if not skip comparing the 2 lists
        if row > 1:
            check = True
        else:
            check = False

        # if check is false, we write in all tools stored in the unused_tools list
        # else we check to see if one of the unused tools have been used and then remove
        # it from the excel database
        row = 1

        if not check:
            for tool in unused_tools:
                result.cell(row=row, column=1, value=tool)
                row += 1

        elif check:
            for r_tool in recorded_unused_tools:
                if r_tool not in unused_tools:
                    result.delete_rows(row)
                
                row += 1
    
        database.save(database_path)



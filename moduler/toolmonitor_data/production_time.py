""" This script is going to be timecalculating my machiningtime,
    and outputting the results in a excelfile """

import openpyxl as op

# We need some refrence to the database exceldocument


def production_time(database_path, ordernumber, total_time):
    # checking if the excel file exists or not, if not generate
    if not database_path.is_file():

        workbook = op.Workbook()

        workbook.create_sheet('Raw Data')
        workbook.create_sheet('Results')

        del workbook['Sheet']

        worksheet = workbook['Raw Data']
        worksheet.cell(row=1, column=1, value='Productcode:')
        worksheet.cell(row=1, column=2, value='Time:')

        worksheet.cell(row=2, column=1, value=ordernumber)
        worksheet.cell(row=2, column=2, value=total_time)

        worksheet = workbook['Results']
        worksheet.cell(row=1, column=1, value='Productcode:')
        worksheet.cell(row=1, column=2, value='Total Machining time:')

        workbook.save(database_path)

    else:
        workbook = op.load_workbook(database_path)
        worksheet = workbook['Raw Data']

        row = 2
        count = 0
        while worksheet.cell(row=row, column=1).value is not None:

            if worksheet.cell(row=row, column=1).value == ordernumber:
                worksheet.cell(row=row, column=1, value=ordernumber)
                worksheet.cell(row=row, column=2, value=total_time)
                count = 1

            row += 1

        if count == 0:
            worksheet.cell(row=row, column=1, value=ordernumber)
            worksheet.cell(row=row, column=2, value=total_time)

        workbook.save(database_path)

    ##############################################################################

    # reading in all product numbers in the raw data sheet into a list
    productnumber = []
    row = 2
    while worksheet.cell(row=row, column=1).value is not None:
        raw_product = worksheet.cell(row=row, column=1).value
        productnumber.append(raw_product[:6])
        row += 1

    # coverting from a list to a set and back to a list to remove duplicates
    productnumber = list(set(productnumber))

    # main loop for finding all entries with the same number, add up their time
    # and write it out to the results sheet

    for item in productnumber:

        time = 0
        row = 2
        while worksheet.cell(row=row, column=1).value is not None:
            ordernumber = worksheet.cell(row=row, column=1).value
            ordernumber = ordernumber[:6]

            if ordernumber == item:
                time += worksheet.cell(row=row, column=2).value

            row += 1

        # switching to the results sheet for summary writing
        worksheet = workbook['Results']

        seconds = 0
        minutes = 0
        hours = 0

        # converting seconds to hours, minutes and seconds
        while time != 0:

            if time > 3600:
                hours += 1
                time -= 3600

            elif time > 60:
                minutes += 1
                time -= 60

            else:
                seconds = time
                time -= time

        converted_total_time = str(hours) + ':' + str(minutes) + ':' + str(seconds)

        # writing the results to the sheet, if number already exists, overwrite time
        # in place
        check = 0
        row_results = 2
        while worksheet.cell(row=row_results, column=1).value is not None:

            if item == worksheet.cell(row=row_results, column=1).value:
                worksheet.cell(row=row_results, column=2, value=converted_total_time)
                check = 1
            row_results += 1
        
        if check == 0:

            worksheet.cell(row=row_results, column=1, value=item)
            worksheet.cell(row=row_results, column=2, value=converted_total_time)

        row += 1

        # switching back sheet because we are in a loop after all
        worksheet = workbook['Raw Data']

        workbook.save(database_path)


if __name__ == '__main__':

    production_time()

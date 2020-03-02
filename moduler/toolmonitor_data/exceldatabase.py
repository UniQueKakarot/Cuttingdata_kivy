
""" This file handles the creation and data moving operations on the rawdatabase.xlsx file """

from pathlib import Path
import configparser

import openpyxl as op

from moduler.toolmonitor_data.tooltable_formatter_a66 import Formatter


class Database:

    def __init__(self, config_path):

        self.config = configparser.ConfigParser()
        self.config.read(config_path)

        self.database = Path(self.config['Paths']['Rawdatabase'])
        self.formatted_file = Path(self.config['Paths']['Fdata'])

        if Path(self.config['Paths']['Rawdata']).is_file():
            self.raw_file = Path(self.config['Paths']['Rawdata'])

        else:
            self.raw_file = Path('./moduler/toolmonitor_data/rawdata/1000')

        self.unused_tools = []
        self.used_tools = []
        self.special_tools = []

        # We should consider putting this as an config option?
        self.column_header = ['Pot Number', 'Tool Number', 'ITN', 'Tool Life', 'Tool Life Remain',
                              'Tool Length', 'Tool Radius', 'Not Sure', 'Alarm State', 'Spindel Load Limit',
                              'Not Sure', 'Kind', 'Not Sure', 'Not Sure']

        # Check if the excel database file exists, if not create it and fill
        # the first worksheet with data

        if not self.database.is_file():
            workbook = op.Workbook()

            worksheet1 = workbook.active
            worksheet1.title = "New Data"

            # Since this is a new database, loop through the header info and write it to DB
            column = 1
            for header in self.column_header:
                worksheet1.cell(row=1, column=column, value=header)
                column += 1

            # Call the file formatter
            raw_data = Formatter(self.raw_file)

            row = 2
            for key in raw_data.tooldata:
                column = 1
                for item in raw_data.tooldata[key]:
                    worksheet1.cell(row=row, column=column, value=item)
                    column += 1

                row += 1

            workbook.save(self.database)

        # Are there any problems running the unused tools methods from here?
        # Should it even be in here?
        # self._tool_usage()
        # self._special_tools()

    def load_new_data(self, new_data):
        
        """ Write in the new tooldata to the excel database """

        workbook = op.load_workbook(self.database)
        
        if 'Old Data' in workbook.sheetnames:

            del workbook['Old Data']
            worksheet = workbook['New Data']
            worksheet.title = "Old Data"
            workbook.create_sheet("New Data")
            worksheet = workbook['New Data']
            self._write_data(worksheet, new_data)

        else:
            worksheet = workbook['New Data']
            worksheet.title = "Old Data"
            workbook.create_sheet("New Data")
            worksheet = workbook['New Data']
            self._write_data(worksheet, new_data)

        workbook.save(self.database)

    def _write_data(self, worksheet, new_data):

        column = 1
        for header in self.column_header:
            worksheet.cell(row=1, column=column, value=header)
            column += 1

        # Call the file formatter
        raw_data = Formatter(new_data)

        row = 2
        for key in raw_data.tooldata:
            column = 1
            for item in raw_data.tooldata[key]:
                worksheet.cell(row=row, column=column, value=item)
                column += 1

            row += 1

    def _tool_usage(self):

        """ Looping through the exceldatabase and comparing new and old data to see if tool life
            has changed """

        workbook = op.load_workbook(self.database)

        if 'Old Data' not in workbook.sheetnames:
            self.unused_tools.append(0)
            return

        worksheet_old = workbook['Old Data']
        worksheet_new = workbook['New Data']

        row = 2
        while worksheet_new.cell(row=row, column=5).value is not None:

            if worksheet_old.cell(row=row, column=5).value == worksheet_new.cell(row=row, column=5).value:
                self.unused_tools.append(worksheet_old.cell(row=row, column=2).value)

            elif worksheet_old.cell(row=row, column=5).value != worksheet_new.cell(row=row, column=5).value:
                self.used_tools.append(worksheet_old.cell(row=row, column=2).value)

            row += 1

    def _special_tools(self):

        """ Loop through the excel database and collect all tools within range 700 - 720 """

        workbook = op.load_workbook(self.database)
        worksheet = workbook['New Data']

        # special tools range
        special_range = {'T' + str(i) for i in range(700, 721, 1)}

        row = 2
        while worksheet.cell(row=row, column=2).value is not None:

            if worksheet.cell(row=row, column=2).value in special_range:
                self.special_tools.append(worksheet.cell(row=row, column=2).value)

            row += 1



